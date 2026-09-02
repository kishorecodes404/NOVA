"""
AI Project Status Report Generator
===================================

Implements the "AI Document Generation with PDF/Excel/DOC support"
flow requested for NOVA:

    User Request -> Collect Data from Existing Agents (PO, Meetings/
    Calendar, Leave, Expense - and Task, once a Task agent exists) ->
    Validate Data -> Apply Template -> Generate PDF/Excel/DOCX ->
    User Review -> Final Document.

Example: "Generate a weekly project status report using my pending
tasks, POs, meetings, and important activities."

This module is a sibling to document_generator.py (which produces a
single-record document - one PO, one leave letter, one expense claim,
one day's meeting minutes). This module instead produces a single
MULTI-SECTION REPORT that rolls up several agents' data over a date
window, in all three requested output formats at once.

FACTS vs WORDING (same split as document_generator.py):
    Every number, id, name, date, and status that appears in the
    report - the pending-PO table, the meeting list, the leave/expense
    tables, the section counts - is pulled straight from NOVA's real
    stores via rag.py in Python. An LLM never invents or edits any of
    that. Only the short Executive Summary paragraph at the top of the
    report is free-form prose, and it is written by an AI model
    (`llm_writer`) using the exact same strict-fact-injection +
    fact-verification approach as document_generator._ai_generate_
    document_content() - the model is only given the real counts to
    describe, and its output is rejected unless every one of those
    counts literally appears in it.

    Unlike document_generator.py's single-record documents (which
    hard-fail via DocumentWriterUnavailable if the writer model can't
    produce verified prose), this report is allowed to still be
    produced end-to-end even if the AI writer is unavailable or its
    output fails verification: it falls back to a deterministic,
    plainly-worded summary paragraph built from the same real counts
    (see `_fallback_summary()`). This is a deliberate difference for
    this document type - a multi-agent weekly report is exactly the
    kind of thing someone wants even when the local model is briefly
    down, and the fallback sentence is just as fact-accurate, only
    less naturally phrased. Every table below the summary is always
    Python-built either way.

TASK AGENT:
    NOVA does not currently have a Task agent/store (no to-do-item
    data source exists in rag.py - only PO, Leave, Expense, Meetings,
    and Mail). Rather than letting an AI invent plausible-looking
    "pending tasks" for a business report - which would be a genuine
    fabrication risk - the report's Tasks section says so plainly and
    is left empty of invented data. Wiring in a real Task agent later
    only requires adding one more `_gather_*_section()` function below
    and passing its result into the three builders' Tasks section.

Layout/visual design for the .docx output reuses document_generator's
shared professional layout helpers (title, metadata table, section
headings, body text, signature block) so a generated report looks
consistent with NOVA's other generated documents. The .pdf and .xlsx
outputs use their own libraries (reportlab / openpyxl) but follow the
same navy-accent, clean-table visual language.

Public entry point: generate_report_evidence(question, conversation_history, llm_writer)
    -> (context: str, sources: list[str], file_info: dict | None, confirmation: str)

file_info, when at least one file was produced, looks like:
    {"files": [{"path": "...", "filename": "...", "format": "docx", "mime": "..."}, ...]}
"""

import os
import uuid
from datetime import datetime, date, timedelta

import rag

from document_generator import (
    DocumentWriterUnavailable,
    GENERATED_DOCS_DIR,
    USER_DISPLAY_NAME,
    _ensure_output_dir,
    _ai_generate_document_content,
    _new_professional_document,
    _add_metadata_block,
    _add_section_heading,
    _add_body_text,
    _add_signature_block,
    _ACCENT_COLOR_HEX,
)

# reportlab (PDF) and openpyxl (Excel) are optional at the module
# level - a missing package should disable that ONE output format
# (with a clear warning surfaced in the confirmation message) rather
# than breaking report generation entirely, since the .docx output
# has no extra dependency beyond what document_generator.py already
# requires.
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_JUSTIFY
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    _REPORTLAB_AVAILABLE = True
except ImportError:
    _REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


_ACCENT_RGB_TUPLE = (0x1F, 0x3A, 0x5F)  # matches document_generator._ACCENT_COLOR

_MIME_TYPES = {
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


# =========================================================
# INTENT DETECTION
# =========================================================

_REPORT_GENERATION_VERBS = (
    "generate", "create", "draft", "prepare", "make", "produce",
    "compile", "put together", "build",
)

# Distinct from document_generator.DOCUMENT_TYPE_SIGNALS (single-
# record PO/leave/expense/meeting-minutes documents) - these phrases
# are specifically about a rolled-up, multi-agent STATUS REPORT.
_REPORT_TYPE_SIGNALS = (
    "status report", "project status report", "project report",
    "weekly report", "weekly status report", "monthly report",
    "monthly status report", "progress report", "activity report",
    "summary report",
)


def looks_like_report_generation_request(question):
    """
    Fast, keyword-based check used by route_query() - True only when
    the message both (a) asks to GENERATE something and (b) names a
    status-report-shaped noun. Checked before document_generator.
    looks_like_document_generation_request() in route_query() so a
    "generate a weekly status report" request is never swallowed by
    that function's generic "a document"/"a copy" fallback.
    """
    q = " " + question.lower().strip() + " "

    has_verb = any(verb in q for verb in _REPORT_GENERATION_VERBS)
    if not has_verb:
        return False

    return any(signal in q for signal in _REPORT_TYPE_SIGNALS)


def _resolve_report_window(question):
    """
    Deterministic (non-AI) resolution of the report's date window.
    Defaults to a 7-day look-back + 7-day look-ahead ("weekly") window
    centered on today, which covers both "what recently happened" and
    "what's coming up" for meetings/POs/leave/expenses. A "month"/
    "monthly" mention widens this to 30 days each way. Plain keyword
    matching - never left to an LLM to decide a date range.
    """
    q = question.lower()
    today = date.today()

    if "month" in q:
        return today - timedelta(days=30), today + timedelta(days=30), "Monthly"

    return today - timedelta(days=7), today + timedelta(days=7), "Weekly"


def _new_report_path(fmt):
    _ensure_output_dir()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"status_report_{stamp}_{uuid.uuid4().hex[:6]}.{fmt}"
    return os.path.join(GENERATED_DOCS_DIR, filename), filename


# =========================================================
# DATA COLLECTION - real records only, straight from rag.py
# =========================================================

def _gather_po_section(user):
    history = rag.get_po_history(user, include_cancelled=False)
    pending = [r for r in history if r.get("status") == "pending"]
    return {
        "pending": pending,
        "recent": history[:10],
        "count_pending": len(pending),
        "count_total": len(history),
    }


def _gather_meetings_section(user, start, end):
    events = rag.get_events_in_range(start, end, user=user)
    return {"events": events, "count": len(events)}


def _gather_leave_section(user):
    history = rag.get_leave_history(user, include_cancelled=False)
    pending = [r for r in history if r.get("status") == "pending"]
    return {
        "pending": pending,
        "recent": history[:10],
        "count_pending": len(pending),
        "count_total": len(history),
    }


def _gather_expense_section(user):
    history = rag.get_expense_history(user, include_cancelled=False)
    pending = [r for r in history if r.get("status") == "pending"]
    return {
        "pending": pending,
        "recent": history[:10],
        "count_pending": len(pending),
        "count_total": len(history),
    }


def _gather_task_section():
    """
    No Task agent/store currently exists in NOVA - see the module
    docstring's "TASK AGENT" note. Returns a clearly-marked
    "unavailable" section rather than any invented task data, so
    every caller (docx/pdf/xlsx builders and the confirmation
    message) can render an honest "not available" note instead of
    silently omitting the section or fabricating content.
    """
    return {"available": False, "items": [], "count": 0}


# =========================================================
# VALIDATION
# =========================================================

def _validate_sections(po, meetings, leave, expense, tasks):
    """
    Deterministic validation of the collected data before the
    template is applied - never an LLM's judgment call. Returns
    (has_any_data, warnings). A report with zero data across every
    section is refused (nothing meaningful to generate); a report
    with SOME data but empty sections elsewhere still proceeds, with
    each empty section's warning surfaced in the final confirmation
    message and rendered as a plain "none found" line in the report
    itself (never invented content).
    """
    warnings = []

    if po["count_total"] == 0:
        warnings.append("no purchase order records were found")
    if meetings["count"] == 0:
        warnings.append("no calendar events were found in the report window")
    if leave["count_total"] == 0:
        warnings.append("no leave records were found")
    if expense["count_total"] == 0:
        warnings.append("no expense records were found")
    if not tasks["available"]:
        warnings.append(
            "no Task agent is currently configured in NOVA, so the "
            "Tasks section could not be populated from real data"
        )

    has_any_data = any([
        po["count_total"], meetings["count"],
        leave["count_total"], expense["count_total"],
    ])

    return has_any_data, warnings


# =========================================================
# EXECUTIVE SUMMARY (AI-authored where possible, verified;
# deterministic fallback otherwise - see module docstring)
# =========================================================

def _fallback_summary(window_label, start, end, po, meetings, leave, expense):
    return (
        f"This {window_label.lower()} status report covers "
        f"{start.strftime('%B %d, %Y')} to {end.strftime('%B %d, %Y')}. "
        f"During this period there are {po['count_pending']} pending "
        f"purchase order(s), {meetings['count']} calendar event(s), "
        f"{leave['count_pending']} pending leave request(s), and "
        f"{expense['count_pending']} pending expense claim(s) on record."
    )


def _generate_summary(llm_writer, window_label, start, end, po, meetings, leave, expense):
    """
    Returns (summary_text, summary_source_label). Tries the AI writer
    first (facts injected exactly, verified - see document_generator.
    _ai_generate_document_content()); falls back to a deterministic
    sentence built from the same real counts if the writer is
    unavailable or its output doesn't verify. Never fails outright -
    see the module docstring for why this report intentionally
    differs from document_generator.py's fail-fast documents.
    """
    required_facts = [
        ("pending purchase orders", str(po["count_pending"])),
        ("calendar events", str(meetings["count"])),
        ("pending leave requests", str(leave["count_pending"])),
        ("pending expense claims", str(expense["count_pending"])),
    ]
    instruction = (
        f"Write a short executive summary (2-3 sentences) opening a "
        f"formal {window_label} Project Status Report covering "
        f"{start.strftime('%B %d, %Y')} to {end.strftime('%B %d, %Y')}. "
        f"State that there are {po['count_pending']} pending purchase "
        f"order(s), {meetings['count']} calendar event(s) in this "
        f"window, {leave['count_pending']} pending leave request(s), "
        f"and {expense['count_pending']} pending expense claim(s). "
        f"The detailed tables follow separately, so do not list "
        f"individual items by name or restate exact monetary amounts."
    )

    if llm_writer is not None:
        try:
            summary = _ai_generate_document_content(
                llm_writer, "Project Status Report", instruction, required_facts
            )
        except Exception:
            summary = None
        if summary:
            return summary, "AI-written (fact-verified)"

    return (
        _fallback_summary(window_label, start, end, po, meetings, leave, expense),
        "deterministic fallback (AI writer unavailable or unverified)",
    )


# =========================================================
# DOCX BUILDER (reuses document_generator.py's shared layout)
# =========================================================

def _build_report_docx(window_label, start, end, summary, po, meetings, leave, expense, tasks):
    path, filename = _new_report_path("docx")

    doc = _new_professional_document(
        f"{window_label} Project Status Report",
        subtitle=f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}",
    )

    _add_metadata_block(doc, [
        ("Prepared For", USER_DISPLAY_NAME),
        ("Report Period", f"{start.strftime('%b %d, %Y')} - {end.strftime('%b %d, %Y')}"),
        ("Generated On", datetime.now().strftime("%B %d, %Y %H:%M")),
    ])

    _add_body_text(doc, summary)

    _add_section_heading(doc, "Pending Purchase Orders")
    if po["pending"]:
        table = doc.add_table(rows=1, cols=4)
        table.style = "Light Grid Accent 1"
        hdr = table.rows[0].cells
        hdr[0].text, hdr[1].text, hdr[2].text, hdr[3].text = (
            "PO ID", "Vendor", "Department", "Total (₹)"
        )
        for record in po["pending"]:
            row = table.add_row().cells
            row[0].text = str(record.get("id", ""))[:8].upper()
            row[1].text = str(record.get("vendor", ""))
            row[2].text = str(record.get("department", ""))
            row[3].text = f"{record.get('total_amount', 0):,.2f}"
    else:
        doc.add_paragraph("No pending purchase orders in this period.")

    _add_section_heading(doc, "Meetings & Calendar Activity")
    if meetings["events"]:
        for label in meetings["events"]:
            doc.add_paragraph(label, style="List Bullet")
    else:
        doc.add_paragraph("No calendar events found in this period.")

    _add_section_heading(doc, "Pending Leave Requests")
    if leave["pending"]:
        table = doc.add_table(rows=1, cols=4)
        table.style = "Light Grid Accent 1"
        hdr = table.rows[0].cells
        hdr[0].text, hdr[1].text, hdr[2].text, hdr[3].text = (
            "Type", "Start", "End", "Status"
        )
        for record in leave["pending"]:
            row = table.add_row().cells
            row[0].text = str(record.get("leave_type", "")).title()
            row[1].text = str(record.get("start", ""))
            row[2].text = str(record.get("end", ""))
            row[3].text = str(record.get("status", ""))
    else:
        doc.add_paragraph("No pending leave requests in this period.")

    _add_section_heading(doc, "Pending Expense Claims")
    if expense["pending"]:
        table = doc.add_table(rows=1, cols=4)
        table.style = "Light Grid Accent 1"
        hdr = table.rows[0].cells
        hdr[0].text, hdr[1].text, hdr[2].text, hdr[3].text = (
            "Category", "Vendor", "Amount (₹)", "Status"
        )
        for record in expense["pending"]:
            row = table.add_row().cells
            row[0].text = str(record.get("category", "")).replace("_", " ").title()
            row[1].text = str(record.get("vendor", "") or "-")
            row[2].text = f"{record.get('amount', 0):,.2f}"
            row[3].text = str(record.get("status", ""))
    else:
        doc.add_paragraph("No pending expense claims in this period.")

    _add_section_heading(doc, "Tasks")
    doc.add_paragraph(
        "No Task agent is currently configured in NOVA, so this "
        "section could not be populated from real task data."
    )

    _add_signature_block(doc, USER_DISPLAY_NAME, closing="Prepared by,")

    doc.save(path)
    return path, filename


# =========================================================
# PDF BUILDER (reportlab) - same sections, matching accent color
# =========================================================

def _build_report_pdf(window_label, start, end, summary, po, meetings, leave, expense, tasks):
    if not _REPORTLAB_AVAILABLE:
        return None

    path, filename = _new_report_path("pdf")
    accent = colors.HexColor(f"#{_ACCENT_COLOR_HEX}")
    muted = colors.HexColor("#5A5A5A")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], textColor=accent,
        fontSize=22, spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"], textColor=muted,
        fontSize=10.5, spaceAfter=10, italic=True,
    )
    heading_style = ParagraphStyle(
        "SectionHeading", parent=styles["Heading2"], textColor=accent,
        fontSize=12, spaceBefore=14, spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"], alignment=TA_JUSTIFY,
        fontSize=11, leading=15, spaceAfter=10,
    )
    normal_style = styles["Normal"]

    def _table(headers, rows):
        data = [headers] + rows
        t = Table(data, hAlign="LEFT", colWidths=None)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), accent),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F9")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    elements = [
        Paragraph(f"{window_label} Project Status Report", title_style),
        Paragraph(
            f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}",
            subtitle_style,
        ),
        Paragraph(
            f"<b>Prepared For:</b> {USER_DISPLAY_NAME} &nbsp;&nbsp; "
            f"<b>Generated On:</b> {datetime.now().strftime('%B %d, %Y %H:%M')}",
            normal_style,
        ),
        Spacer(1, 10),
        Paragraph(summary, body_style),
        Paragraph("PENDING PURCHASE ORDERS", heading_style),
    ]

    if po["pending"]:
        rows = [
            [
                str(r.get("id", ""))[:8].upper(),
                str(r.get("vendor", "")),
                str(r.get("department", "")),
                f"{r.get('total_amount', 0):,.2f}",
            ]
            for r in po["pending"]
        ]
        elements.append(_table(["PO ID", "Vendor", "Department", "Total (₹)"], rows))
    else:
        elements.append(Paragraph("No pending purchase orders in this period.", normal_style))

    elements.append(Paragraph("MEETINGS &amp; CALENDAR ACTIVITY", heading_style))
    if meetings["events"]:
        for label in meetings["events"]:
            elements.append(Paragraph(f"&bull; {label}", normal_style))
    else:
        elements.append(Paragraph("No calendar events found in this period.", normal_style))

    elements.append(Paragraph("PENDING LEAVE REQUESTS", heading_style))
    if leave["pending"]:
        rows = [
            [
                str(r.get("leave_type", "")).title(),
                str(r.get("start", "")),
                str(r.get("end", "")),
                str(r.get("status", "")),
            ]
            for r in leave["pending"]
        ]
        elements.append(_table(["Type", "Start", "End", "Status"], rows))
    else:
        elements.append(Paragraph("No pending leave requests in this period.", normal_style))

    elements.append(Paragraph("PENDING EXPENSE CLAIMS", heading_style))
    if expense["pending"]:
        rows = [
            [
                str(r.get("category", "")).replace("_", " ").title(),
                str(r.get("vendor", "") or "-"),
                f"{r.get('amount', 0):,.2f}",
                str(r.get("status", "")),
            ]
            for r in expense["pending"]
        ]
        elements.append(_table(["Category", "Vendor", "Amount (₹)", "Status"], rows))
    else:
        elements.append(Paragraph("No pending expense claims in this period.", normal_style))

    elements.append(Paragraph("TASKS", heading_style))
    elements.append(Paragraph(
        "No Task agent is currently configured in NOVA, so this "
        "section could not be populated from real task data.",
        normal_style,
    ))

    doc = SimpleDocTemplate(
        path, pagesize=A4,
        topMargin=0.6 * inch, bottomMargin=0.75 * inch,
        leftMargin=0.9 * inch, rightMargin=0.9 * inch,
    )
    doc.build(elements)

    return path, filename


# =========================================================
# XLSX BUILDER (openpyxl) - one sheet per section + a Summary sheet
# =========================================================

def _build_report_xlsx(window_label, start, end, summary, po, meetings, leave, expense, tasks):
    if not _OPENPYXL_AVAILABLE:
        return None

    path, filename = _new_report_path("xlsx")

    accent_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10.5)
    title_font = Font(bold=True, color="1F3A5F", size=16)
    subtitle_font = Font(italic=True, color="5A5A5A", size=10.5)
    wrap_center = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

    def _style_header_row(ws, row_idx, num_cols):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = accent_fill
            cell.font = header_font
            cell.alignment = wrap_center
            cell.border = thin_border

    def _autosize(ws, num_cols, min_width=12, max_width=45):
        for col_idx in range(1, num_cols + 1):
            letter = get_column_letter(col_idx)
            longest = max(
                (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
                default=min_width,
            )
            ws.column_dimensions[letter].width = min(max(longest + 2, min_width), max_width)

    def _write_table(ws, start_row, headers, rows, empty_message):
        ws.cell(row=start_row, column=1, value=None)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col_idx, value=header)
        _style_header_row(ws, start_row, len(headers))

        if rows:
            for r_offset, row_values in enumerate(rows, start=1):
                for col_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=start_row + r_offset, column=col_idx, value=value)
                    cell.border = thin_border
        else:
            ws.cell(row=start_row + 1, column=1, value=empty_message)

        _autosize(ws, len(headers))

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{window_label} Project Status Report"
    ws["A1"].font = title_font
    ws["A2"] = f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}"
    ws["A2"].font = subtitle_font
    ws["A4"] = "Prepared For"
    ws["B4"] = USER_DISPLAY_NAME
    ws["A5"] = "Generated On"
    ws["B5"] = datetime.now().strftime("%B %d, %Y %H:%M")
    ws["A4"].font = Font(bold=True)
    ws["A5"].font = Font(bold=True)
    ws["A7"] = "Executive Summary"
    ws["A7"].font = Font(bold=True, color="1F3A5F", size=12)
    ws["A8"] = summary
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[8].height = 60
    ws.merge_cells("A8:D8")

    ws["A10"] = "Section"
    ws["B10"] = "Count"
    _style_header_row(ws, 10, 2)
    counts = [
        ("Pending Purchase Orders", po["count_pending"]),
        ("Calendar Events", meetings["count"]),
        ("Pending Leave Requests", leave["count_pending"]),
        ("Pending Expense Claims", expense["count_pending"]),
        ("Tasks (no Task agent configured)", "N/A"),
    ]
    for offset, (label, value) in enumerate(counts, start=1):
        ws.cell(row=10 + offset, column=1, value=label)
        ws.cell(row=10 + offset, column=2, value=value)
    _autosize(ws, 2)

    # ---- Purchase Orders sheet ----
    ws_po = wb.create_sheet("Purchase Orders")
    _write_table(
        ws_po, 1,
        ["PO ID", "Vendor", "Department", "Status", "Total (₹)"],
        [
            [
                str(r.get("id", ""))[:8].upper(),
                r.get("vendor", ""),
                r.get("department", ""),
                r.get("status", ""),
                r.get("total_amount", 0),
            ]
            for r in po["pending"]
        ],
        "No pending purchase orders in this period.",
    )

    # ---- Meetings sheet ----
    ws_meet = wb.create_sheet("Meetings")
    _write_table(
        ws_meet, 1,
        ["Event"],
        [[label] for label in meetings["events"]],
        "No calendar events found in this period.",
    )

    # ---- Leave sheet ----
    ws_leave = wb.create_sheet("Leave Requests")
    _write_table(
        ws_leave, 1,
        ["Type", "Start", "End", "Status"],
        [
            [
                str(r.get("leave_type", "")).title(),
                r.get("start", ""),
                r.get("end", ""),
                r.get("status", ""),
            ]
            for r in leave["pending"]
        ],
        "No pending leave requests in this period.",
    )

    # ---- Expenses sheet ----
    ws_exp = wb.create_sheet("Expense Claims")
    _write_table(
        ws_exp, 1,
        ["Category", "Vendor", "Amount (₹)", "Status"],
        [
            [
                str(r.get("category", "")).replace("_", " ").title(),
                r.get("vendor", "") or "-",
                r.get("amount", 0),
                r.get("status", ""),
            ]
            for r in expense["pending"]
        ],
        "No pending expense claims in this period.",
    )

    # ---- Tasks sheet (explicitly marked unavailable, no invented data) ----
    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks["A1"] = (
        "No Task agent is currently configured in NOVA, so this "
        "sheet could not be populated from real task data."
    )
    ws_tasks.column_dimensions["A"].width = 80

    wb.save(path)
    return path, filename


# =========================================================
# PUBLIC ENTRY POINT
# =========================================================

def generate_report_evidence(question, conversation_history="", llm_writer=None, user="me"):
    """
    Returns (context, sources, file_info, confirmation) - same shape
    as document_generator.generate_document_evidence(), see that
    module's docstring for the general contract. file_info here is
    either None (nothing could be generated) or {"files": [...]}
    listing every format that was actually produced.
    """

    start, end, window_label = _resolve_report_window(question)

    try:
        po = _gather_po_section(user)
        meetings = _gather_meetings_section(user, start, end)
        leave = _gather_leave_section(user)
        expense = _gather_expense_section(user)
        tasks = _gather_task_section()
    except Exception as error:
        message = f"Report generation failed while collecting data: {error}"
        return message, [], None, message

    has_any_data, data_warnings = _validate_sections(po, meetings, leave, expense, tasks)

    if not has_any_data:
        message = (
            f"I couldn't generate a {window_label.lower()} status report - "
            f"no purchase order, meeting, leave, or expense records were "
            f"found for {start.strftime('%B %d, %Y')} to "
            f"{end.strftime('%B %d, %Y')}."
        )
        return message, [], None, message

    summary, summary_source = _generate_summary(
        llm_writer, window_label, start, end, po, meetings, leave, expense
    )

    build_args = (window_label, start, end, summary, po, meetings, leave, expense, tasks)

    files = []
    format_warnings = []

    try:
        docx_path, docx_filename = _build_report_docx(*build_args)
        files.append({
            "path": docx_path, "filename": docx_filename,
            "format": "docx", "mime": _MIME_TYPES["docx"],
        })
    except Exception as error:
        format_warnings.append(f"DOCX generation failed: {error}")

    try:
        pdf_result = _build_report_pdf(*build_args)
        if pdf_result:
            pdf_path, pdf_filename = pdf_result
            files.append({
                "path": pdf_path, "filename": pdf_filename,
                "format": "pdf", "mime": _MIME_TYPES["pdf"],
            })
        else:
            format_warnings.append(
                "PDF generation skipped: the 'reportlab' package is not installed."
            )
    except Exception as error:
        format_warnings.append(f"PDF generation failed: {error}")

    try:
        xlsx_result = _build_report_xlsx(*build_args)
        if xlsx_result:
            xlsx_path, xlsx_filename = xlsx_result
            files.append({
                "path": xlsx_path, "filename": xlsx_filename,
                "format": "xlsx", "mime": _MIME_TYPES["xlsx"],
            })
        else:
            format_warnings.append(
                "Excel generation skipped: the 'openpyxl' package is not installed."
            )
    except Exception as error:
        format_warnings.append(f"Excel generation failed: {error}")

    if not files:
        message = (
            "I gathered the report data, but couldn't produce any output "
            "file: " + " ".join(format_warnings)
        )
        return message, [], None, message

    evidence_lines = [
        f"WINDOW: {window_label} ({start} to {end})",
        f"PENDING_PO_COUNT: {po['count_pending']}",
        f"MEETING_COUNT: {meetings['count']}",
        f"PENDING_LEAVE_COUNT: {leave['count_pending']}",
        f"PENDING_EXPENSE_COUNT: {expense['count_pending']}",
        f"TASK_AGENT: not configured",
        f"SUMMARY_SOURCE: {summary_source}",
        f"FORMATS_GENERATED: {', '.join(f['format'].upper() for f in files)}",
    ]
    if data_warnings:
        evidence_lines.append("DATA_WARNINGS: " + "; ".join(data_warnings))
    if format_warnings:
        evidence_lines.append("FORMAT_WARNINGS: " + "; ".join(format_warnings))

    sources = [f"PO {r.get('id', '')} - {r.get('vendor', '')}" for r in po["pending"]]
    sources += [f"Calendar event - {label}" for label in meetings["events"]]
    sources += [
        f"Leave request {r.get('id', '')} - {r.get('leave_type', '')} "
        f"({r.get('start', '')} to {r.get('end', '')})"
        for r in leave["pending"]
    ]
    sources += [
        f"Expense claim {r.get('id', '')} - {r.get('category', '')} "
        f"(₹{r.get('amount', 0):,.2f})"
        for r in expense["pending"]
    ]

    context = (
        f"A {window_label} Project Status Report was generated for "
        f"{start.strftime('%B %d, %Y')} to {end.strftime('%B %d, %Y')}, "
        f"available as: {', '.join(f['filename'] for f in files)}. "
        f"It was built from the following real, verified data:\n\n"
        + "\n".join(evidence_lines)
    )

    formats_label = " / ".join(f["format"].upper() for f in files)
    confirmation = (
        f"Your **{window_label} Project Status Report** is ready "
        f"({formats_label}), covering **{start.strftime('%b %d, %Y')} - "
        f"{end.strftime('%b %d, %Y')}**: **{po['count_pending']}** pending "
        f"PO(s), **{meetings['count']}** calendar event(s), "
        f"**{leave['count_pending']}** pending leave request(s), and "
        f"**{expense['count_pending']}** pending expense claim(s)."
    )
    if data_warnings:
        confirmation += (
            " Note: " + "; ".join(data_warnings).capitalize() + "."
        )
    if format_warnings:
        confirmation += " " + " ".join(format_warnings)

    return context, sources, {"files": files}, confirmation
